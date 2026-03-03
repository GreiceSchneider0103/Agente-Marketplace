import os
import json
import logging
import re
import io
import csv
from typing import Any, Dict, Tuple

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for

from docx_builder import generate_docx_from_data
from marketplace_specialist import MarketplaceSpecialist

# Optional deps (upload parsing)
try:
    import pdfplumber  # type: ignore
except Exception:
    pdfplumber = None

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)

# Session secret (support both names)
app.secret_key = os.getenv("SECRET_KEY") or os.getenv("FLASK_SECRET_KEY") or "dev-secret-key"

# Instantiate specialist
specialist = MarketplaceSpecialist()


def login_required(fn):
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)

    wrapper.__name__ = fn.__name__
    return wrapper


# ----------------------------
# Upload parsing (PDF/Sheet)
# ----------------------------

def _extract_structured_from_text(text: str) -> Dict[str, str]:
    """Best-effort extraction of fields from text."""
    if not text:
        return {}

    def _first(patterns):
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                return m.group(1).strip()
        return ""

    out: Dict[str, str] = {}

    nome = _first([
        r"(?:produto|nome do produto)\s*[:\-]\s*([^\n\r]+)",
    ])
    if nome:
        out["nome_produto"] = nome

    marca = _first([
        r"(?:marca|linha)\s*[:\-]\s*([^\n\r]+)",
        r"fabricante\s*[:\-]\s*([^\n\r]+)",
    ])
    if marca:
        out["marca_linha"] = marca

    materiais = _first([
        r"materiais?\s*[:\-]\s*([^\n\r]+)",
        r"estrutura\s*[:\-]\s*([^\n\r]+)",
    ])
    if materiais:
        out["materiais"] = materiais

    dim = _first([
        r"dimens(?:o|õ)es\s*[:\-]\s*([\d,.]+\s*[xX]\s*[\d,.]+\s*[xX]\s*[\d,.]+\s*(?:cm|mm|m)?)",
        r"([\d,.]+\s*[xX]\s*[\d,.]+\s*[xX]\s*[\d,.]+\s*(?:cm|mm|m))",
    ])
    if dim:
        out["dimensoes"] = dim

    peso = _first([
        r"peso\s*suportado\s*[:\-]\s*([\d,.]+\s*kg)",
        r"suporta\s*at(?:e|é)\s*([\d,.]+\s*kg)",
    ])
    if peso:
        out["peso_suportado"] = peso

    conteudo = _first([
        r"conte(?:u|ú)do\s+da\s+embalagem\s*[:\-]\s*([^\n\r]+)",
        r"itens\s+inclusos\s*[:\-]\s*([^\n\r]+)",
    ])
    if conteudo:
        out["conteudo_embalagem"] = conteudo

    montagem = _first([
        r"necessita\s+montagem\s*[:\-]\s*(sim|n(?:a|ã)o)",
        r"montagem\s*[:\-]\s*(sim|n(?:a|ã)o)",
    ]).lower()
    if montagem in {"sim", "não", "nao"}:
        out["necessita_montagem"] = "sim" if montagem == "sim" else "nao"

    return out


def _read_pdf_text(file_storage) -> str:
    if not pdfplumber:
        return ""
    try:
        file_storage.stream.seek(0)
        with pdfplumber.open(file_storage.stream) as pdf:
            parts = []
            for page in pdf.pages[:10]:
                t = page.extract_text() or ""
                if t:
                    parts.append(t)
        return "\n".join(parts)
    except Exception as e:
        app.logger.warning(f"PDF parse failed: {e}")
        return ""


def _read_sheet_preview(file_storage) -> str:
    filename = (file_storage.filename or "").lower()
    try:
        file_storage.stream.seek(0)
        data = file_storage.stream.read()
    except Exception:
        return ""

    if filename.endswith(".csv"):
        try:
            s = data.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(s))
            rows = []
            for i, row in enumerate(reader):
                rows.append("\t".join(row))
                if i >= 40:
                    break
            return "\n".join(rows)
        except Exception:
            return ""

    if filename.endswith(".xlsx") and load_workbook:
        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append("\t".join(["" if v is None else str(v) for v in row]))
                if r_idx >= 40:
                    break
            return "\n".join(rows)
        except Exception as e:
            app.logger.warning(f"XLSX parse failed: {e}")
            return ""

    return ""


def process_uploads(files) -> Tuple[Dict[str, str], Dict[str, Any]]:
    """Return (extracted_fields, uploads_context)."""
    extracted: Dict[str, str] = {}
    uploads_ctx: Dict[str, Any] = {}

    pdf = files.get("file_pdf")
    if pdf and getattr(pdf, "filename", ""):
        pdf_text = _read_pdf_text(pdf)
        uploads_ctx["pdf_name"] = pdf.filename
        uploads_ctx["pdf_text"] = pdf_text[:8000] if pdf_text else ""
        if pdf_text:
            extracted.update(_extract_structured_from_text(pdf_text))

    sheet = files.get("file_sheet")
    if sheet and getattr(sheet, "filename", ""):
        preview = _read_sheet_preview(sheet)
        uploads_ctx["sheet_name"] = sheet.filename
        uploads_ctx["sheet_preview"] = preview[:8000] if preview else ""
        if preview:
            extracted.update(_extract_structured_from_text(preview))

    photo = files.get("file_photo")
    if photo and getattr(photo, "filename", ""):
        uploads_ctx["photo_name"] = photo.filename

    return extracted, uploads_ctx


@app.get("/healthz")
def healthz():
    return "OK", 200


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")
        if email == os.getenv("ADMIN_EMAIL") and password == os.getenv("ADMIN_PASSWORD"):
            session["logged_in"] = True
            return redirect(url_for("home"))
        return render_template("login.html", error="Credenciais inválidas")
    return render_template("login.html")


@app.get("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect(url_for("login"))


@app.get("/")
@login_required
def home():
    return render_template("index.html")


@app.post("/generate")
@login_required
def generate():
    """
    Recebe multipart/form-data (payload JSON + uploads) ou JSON puro.
    IMPORTANTE: seu script.js envia FormData com:
      - payload: JSON.stringify({ product_data: state.data, meta: state.meta })
      - file_pdf / file_photo / file_sheet (opcionais)
    """
    try:
        if request.is_json:
            payload = request.get_json(silent=True) or {}
        else:
            payload_raw = request.form.get("payload", "{}")
            payload = json.loads(payload_raw or "{}")

        product_data = payload.get("product_data") or {}
        meta = payload.get("meta") or {}

        extracted, uploads_ctx = process_uploads(request.files)
        merged_product = {**extracted, **product_data}  # manual vence extraído

        marketplace = merged_product.get("marketplace_alvo") or merged_product.get("marketplace")

        result = specialist.run(
            product_data=merged_product,
            marketplace=marketplace,
            uploads={"context": uploads_ctx, "meta": meta},
        )

        return jsonify(result)
    except Exception as e:
        app.logger.exception("/generate failed")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.post("/generate-docx")
@login_required
def generate_docx():
    try:
        if request.is_json:
            payload = request.get_json(silent=True) or {}
        else:
            payload_raw = request.form.get("payload", "{}")
            payload = json.loads(payload_raw or "{}")

        product_data = payload.get("product_data") or {}
        meta = payload.get("meta") or {}

        extracted, uploads_ctx = process_uploads(request.files)
        merged_product = {**extracted, **product_data}
        marketplace = merged_product.get("marketplace_alvo") or merged_product.get("marketplace")

        result = specialist.run(
            product_data=merged_product,
            marketplace=marketplace,
            uploads={"context": uploads_ctx, "meta": meta},
        )

        if result.get("status") != "ok":
            # contrato da UI: missing_fields também precisa ser 200
            return jsonify(result), 200

        # manter contrato do seu docx_builder: retorna file-like (BytesIO)
        docx_file = generate_docx_from_data(result)
        return send_file(
            docx_file,
            as_attachment=True,
            download_name="anuncio.docx",
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
    except Exception as e:
        app.logger.exception("/generate-docx failed")
        return jsonify({"status": "error", "message": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=True)